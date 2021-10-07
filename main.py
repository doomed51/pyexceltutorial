#load the workbook
#select sheet by name 
# 
#   useful commands
#
# sheet.title
# sheet["A1"].value
# sheet.cell(row=10, column=6).value
# sheet["C1"] = "writing ;)"
# workbook.save(filename="hello_world_append.xlsx")
#
# More here: https://realpython.com/openpyxl-excel-spreadsheets-python/
#
#   sheet[sheet.max_row]

from openpyxl import workbook, load_workbook
from openpyxl.worksheet import worksheet

#load the workbook
workbook = load_workbook(filename="Sample.xlsm")

#select sheet by name 
sheet = workbook["Valuations"]

for row in sheet.iter_rows(values_only=True):
    print(row)